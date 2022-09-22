from openpyxl import load_workbook, Workbook
import os


def get_codes(filename):
    # получаю адреса ячеек с непустым штрих кодом
    wb = load_workbook(filename=filename)
    page = wb.active
    data = []
    for i in page.iter_rows(min_row=2, min_col=4):
        if (i[0].value) is not None:
            data.append((page.cell(row=i[0].row, column=2).value, i[0].row))
    return data


def generate_photos(data, filename):
    for i in data:
        copy(filename, i[0])


def copy(name, new_name):
    # более правильно, на мой взгляд было бы использовать модуль shutil, так как он не имеет
    # различий в синтаксисе для linux и windows. Но я использовал указанный в задании модуль os
    if os.name == 'nt':
        os.popen(f'copy {name} files\{new_name}.jpg')
    elif os.name == 'posix':
        os.popen(f'cp {name} files/{new_name}.jpg')
    else:
        raise OSError


def create_links(filename, data):
    # здесь я создаю  колонку фото, где находятся гиперссылки на фото, с текстом, дублирующим
    # колонку адресс ячейки, как и было написано в задании
    wb = load_workbook(filename=filename)
    page = wb.active
    page['E1'] = 'Фото'
    for i in data:
        page.cell(row=i[1], column=5).value = i[0]
        page.cell(row=i[1], column=5).hyperlink = f'files/{i[0]}.jpg'
    wb.save(filename)


data = get_codes('fixed_file.xlsx')
generate_photos(data, 'cat.jpg')
create_links('fixed_file.xlsx', data)
