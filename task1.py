from openpyxl import load_workbook, Workbook


def get_data(location, cells_count):
    # прочитал ячейки таблицы, переместил значения в список

    # из задания не совсем понятно - ячейки уже идут в виде змейки, тогда,
    # их уже не надо менять местами и нужно присвоить им просто номер (за 69-1 идет 69-2, потом 68-2)
    # или же если 70 ячейка в таблице - находится над ячейкой 1, тогда
    # мне надо поменять порядок прохождения ячеек в четном слое, и присвоить им номера как показано выше
    wb = load_workbook(location)["Лист1"]
    layers = []
    count = 0
    layer = []
    for row in wb.values:
        if len(layer) < cells_count:
            # if count % 2 == 0:
            layer.append(row)
            # else:
            #    layer.insert(0, row)
        else:
            layers.append(layer)
            layer = []
            layer.append(row)
            count += 1
    layers.append(layer)
    return layers


def generate_address(layers, cells_count):
    # присваиваю адрес каждому штрих коду
    addresses = []
    layer_count = 0
    for i in layers:
        layer_count += 1
        if layers.index(i) % 2 == 0:
            element_count = 0
            for j in i:
                element_count += 1
                addresses.append((j[0], f'{element_count}-{layer_count}'))
        else:
            element_count = cells_count + 1
            for j in i:
                element_count -= 1
                addresses.append((j[0], f'{element_count}-{layer_count}'))
    return addresses


def write_data(title, addresses):
    # запись данных в файл
    wb = Workbook()
    page = wb.active
    page.title = 'Data'
    row_count = 0
    page['A1'] = '№'
    page['B1'] = 'Адрес'
    page['C1'] = 'Штрих-код'
    page['D1'] = 'Отредактированный Штрих-код'
    for i in addresses:
        page.cell(column=1, row=row_count + 2, value=addresses.index(i) + 1)
        page.cell(column=2, row=row_count + 2, value=i[1])
        changed = i[0]
        if '00' in changed:
            changed = changed[2:]
            page.cell(column=3, row=row_count + 2, value=i[0])
            page.cell(column=4, row=row_count + 2, value=changed)
        row_count += 1
    wb.save(filename=title)


layers = get_data('tz_data.xlsx', 69)
addresses = generate_address(layers, 69)
write_data('fixed_file.xlsx', addresses)
